# gui/export_tab.py

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QPushButton, QProgressBar,
    QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from backend.preservica_client import PreservicaClient
import pyPreservica as pyp
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class ExportWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)

    def __init__(self, client, ref_list, export_path):
        super().__init__()
        self.client = client
        self.ref_list = ref_list
        self.export_path = export_path

    def run(self):
        rows = []
        fieldnames = {"reference", "title", "type", "qdc_xml"}

        total = len(self.ref_list)
        for i, ref in enumerate(self.ref_list, 1):
            try:
                entity = self.client.asset(ref)
                etype = "ASSET"
            except Exception:
                try:
                    entity = self.client.folder(ref)
                    etype = "FOLDER"
                except Exception:
                    continue  # skip invalid refs

            row = {
                "reference": entity.reference,
                "title": entity.title,
                "type": etype,
                "qdc_xml": ""
            }

            for url, schema in (entity.metadata or {}).items():
                if "dc" in schema.lower():
                    try:
                        xml = self.client.metadata(url).strip()
                        row["qdc_xml"] = xml
                        root = ET.fromstring(xml)
                        ns = {
                            "dc": "http://purl.org/dc/elements/1.1/",
                            "dcterms": "http://purl.org/dc/terms/"
                        }
                        counts = {}
                        for prefix in ns:
                            for elem in root.findall(f".//{{{ns[prefix]}}}*"):
                                tag = elem.tag.split("}")[-1]
                                value = (elem.text or "").strip()
                                if not value:
                                    continue
                                base = f"dc:{tag}"
                                count = counts.get(base, 0)
                                col = base if count == 0 else f"{base}.{count}"
                                row[col] = value
                                fieldnames.add(col)
                                counts[base] = count + 1
                    except ET.ParseError:
                        continue

            rows.append(row)
            self.progress.emit(int(i / total * 100))

        # Write to Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Metadata"

        headers = sorted(fieldnames - {"reference", "title", "type"})
        final_headers = ["reference", "title", "type"] + headers
        ws.append(final_headers)

        for i, header in enumerate(final_headers, 1):
            col_letter = get_column_letter(i)
            ws[f"{col_letter}1"].font = Font(bold=True)

        for row_data in rows:
            row = [row_data.get(h, "") for h in final_headers]
            ws.append(row)

        wb.save(self.export_path)
        self.finished.emit(self.export_path)


class ExportTab(QWidget):
    def __init__(self, client):
        super().__init__()
        self.client = client

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.status_label = QLabel("Select assets or folders to export metadata.")
        self.layout.addWidget(self.status_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.layout.addWidget(self.progress_bar)

        self.export_button = QPushButton("Start Export")
        self.export_button.clicked.connect(self.start_export)
        self.layout.addWidget(self.export_button)

        self.ref_list = []

        self.worker = None

    def start_export(self):
        if not self.ref_list:
            QMessageBox.warning(self, "No Items", "No items selected for export.")
            return

        export_path, _ = QFileDialog.getSaveFileName(self, "Save Metadata", filter="Excel Files (*.xlsx)")
        if not export_path:
            return
        if not export_path.endswith(".xlsx"):
            export_path += ".xlsx"

        self.start_export_with_refs(self.ref_list, export_path)

    def start_export_with_refs(self, ref_list, export_path):
        self.ref_list = ref_list
        self.progress_bar.setValue(0)
        self.status_label.setText("Exporting metadata...")

        self.worker = ExportWorker(self.client, ref_list, export_path)
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.finished.connect(self.export_finished)
        self.worker.start()

    def export_finished(self, export_path):
        self.status_label.setText("Export complete!")
        QMessageBox.information(self, "Export Complete", f"Metadata exported to:\n{export_path}")
        self.progress_bar.setValue(100)
