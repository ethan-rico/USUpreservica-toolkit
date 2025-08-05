# gui/update_tab.py

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QPushButton,
    QProgressBar, QFileDialog, QMessageBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from backend.preservica_client import PreservicaClient
import openpyxl
import xml.etree.ElementTree as ET
from xml.dom import minidom
import pyPreservica as pyp


class UpdateWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(int, int)

    def __init__(self, client, file_path):
        super().__init__()
        self.client = client
        self.file_path = file_path

    def run(self):
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
        except Exception:
            self.finished.emit(0, 0)
            return

        headers = [cell.value for cell in ws[1]]
        total = ws.max_row - 1
        updated = 0
        skipped = 0

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            data = dict(zip(headers, row))
            ref = data.get("reference")
            entity_type = data.get("type")

            if not ref or not entity_type:
                skipped += 1
                continue

            try:
                entity = (
                    self.client.asset(ref)
                    if entity_type == "ASSET"
                    else self.client.folder(ref)
                )
            except Exception:
                skipped += 1
                continue

            # Build new QDC XML
            root = ET.Element("oai_qdc:qualifieddc", {
                "xmlns:oai_qdc": "http://www.openarchives.org/OAI/2.0/oai_qdc/",
                "xmlns:dc": "http://purl.org/dc/elements/1.1/",
                "xmlns:dcterms": "http://purl.org/dc/terms/"
            })

            for key, value in data.items():
                if key.startswith("dc:") and value:
                    tag_parts = key.split(":")
                    if len(tag_parts) > 1:
                        base_tag = tag_parts[1].split(".")[0]
                        if base_tag:
                            if "." in key:
                                # Allow for multiple same tags like dc:subject.0, dc:subject.1
                                base_tag = base_tag
                            if "dcterms:" in key:
                                sub_el = ET.SubElement(root, f"{{http://purl.org/dc/terms/}}{base_tag}")
                            else:
                                sub_el = ET.SubElement(root, f"{{http://purl.org/dc/elements/1.1/}}{base_tag}")
                            sub_el.text = str(value)

            raw_xml = ET.tostring(root, encoding="utf-8")
            pretty_xml = minidom.parseString(raw_xml).toprettyxml(indent="  ")

            try:
                self.client.update_metadata(ref, "http://purl.org/dc/terms/", pretty_xml)
                updated += 1
            except Exception:
                skipped += 1

            self.progress.emit(int((i / total) * 100))

        self.finished.emit(updated, skipped)


class UpdateTab(QWidget):
    def __init__(self, client):
        super().__init__()
        self.client = client
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        self.info_label = QLabel("Select an Excel file with updated metadata to apply to Preservica assets.")
        self.layout.addWidget(self.info_label)

        self.update_button = QPushButton("Update Metadata from .xlsx File")
        self.update_button.clicked.connect(self.handle_update)
        self.layout.addWidget(self.update_button)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.layout.addWidget(self.progress_bar)

    def handle_update(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", filter="Excel Files (*.xlsx)")
        if not file_path:
            return

        self.progress_bar.setValue(0)
        self.worker = UpdateWorker(self.client, file_path)
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.finished.connect(self.show_results)
        self.worker.start()

    def show_results(self, updated, skipped):
        msg = f"Updated {updated} item(s)."
        if skipped > 0:
            msg += f"\nSkipped {skipped} item(s)."
        QMessageBox.information(self, "Update Complete", msg)
