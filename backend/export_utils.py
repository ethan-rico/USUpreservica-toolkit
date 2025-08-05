from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def export_to_xlsx(path, rows, fieldnames):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata Export"

    # Write headers
    for col_idx, field in enumerate(fieldnames, start=1):
        ws.cell(row=1, column=col_idx, value=field)

    # Write rows, force cell format to "Text"
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, field in enumerate(fieldnames, start=1):
            val = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = "@"  # This locks the cell as plain text

    # Optional: Set each column format to text explicitly (redundant but safe)
    for col_idx in range(1, len(fieldnames) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].number_format = "@"

    wb.save(path)
    return path
