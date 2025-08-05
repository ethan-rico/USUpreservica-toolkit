# logic/operations.py

import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pyPreservica as pyp


def export_metadata_to_excel(client, refs, export_path, progress_callback=None):
    rows = []
    fieldnames = {"reference", "title", "type", "qdc_xml"}

    for index, (ref, ref_type) in enumerate(refs):
        try:
            entity = client.asset(ref) if ref_type == "ASSET" else client.folder(ref)
            etype = "ASSET" if ref_type == "ASSET" else "FOLDER"
        except Exception:
            continue

        row = {
            "reference": entity.reference,
            "title": entity.title,
            "type": etype,
            "qdc_xml": ""
        }

        for url, schema in (entity.metadata or {}).items():
            if "dc" in schema.lower():
                try:
                    xml = client.metadata(url).strip()
                    row["qdc_xml"] = xml
                    root = ET.fromstring(xml)
                    ns = {
                        "dc": "http://purl.org/dc/elements/1.1/",
                        "dcterms": "http://purl.org/dc/terms/"
                    }
                    counts = {}
                    for prefix in ns:
                        for elem in root.findall(f".//{{{ns[prefix]}}}*"):
                            tag = elem.tag.split("}")[-1]
                            value = (elem.text or "").strip()
                            if not value:
                                continue
                            base = f"dc:{tag}"
                            count = counts.get(base, 0)
                            col = base if count == 0 else f"{base}.{count}"
                            row[col] = value
                            fieldnames.add(col)
                            counts[base] = count + 1
                except ET.ParseError:
                    continue

        rows.append(row)

        if progress_callback:
            progress_callback(index + 1, len(refs))

    # Write to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    headers = sorted(fieldnames - {"reference", "title", "type"})
    final_headers = ["reference", "title", "type"] + headers
    ws.append(final_headers)

    for i, header in enumerate(final_headers, 1):
        col_letter = get_column_letter(i)
        ws[f"{col_letter}1"].font = Font(bold=True)

    for row_data in rows:
        row = [row_data.get(h, "") for h in final_headers]
        ws.append(row)

    wb.save(export_path)
